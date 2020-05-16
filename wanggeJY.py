#!/usr/bin/env phython3
# -*- coding: utf-8 -*-

# 20200406 修正 -- version2
#    原因：2020年1到4月，因新冠疫情导致的极端下跌行情
#    修正：当买入4个档位后，后面买入的档位在前一个的基础上+1.5%
#    说明：第5格买入之后，市场／个股可能已经出现极端情况，从第6格开始更多的防守，保证资金安全为前提

__author__ = "challengezcy@163.com"

# 说明：
# 为了资金安全，触发价通常会选择相对低价，如60日线均价再下调10%～20%
# 这时候，高于触发价的网格买入规则，第二格或者第三格作为第一次手动买入

from openpyxl import Workbook
from sys import argv
import pandas as pd
import os

#price: 第二档买入价
#Trade_mount: 此次网格交易总共投入的实际资金
# load: 相当于加多少杠杆交易（0～100）

# python wanggeJY.py 18.79 200000 10
_, price, Trade_mount, loan = argv

# 默认10个网格
GRID_COUNT = 10
UP_BASE_PRICE_GRID_COUNT = 4

# 默认涨6%卖出
UP_range = 6
up_range = (int)(UP_range)

# 默认跌6%买入，参考20200406修正说明
down_range = up_range

# base_price: 基准买入价，用来计算每个网格的买入量
base_price = ((float)(price)) * ((100 - (int)(down_range)) / 100)**3

# 不加借贷的资金额
Trade_mount = (int)(Trade_mount)
init_mount = Trade_mount

# 加借贷后的资金额
loan = (int)(loan)
Trade_mount = Trade_mount * (100 + loan) / 100

# 以base_price为基准，平均每个grid可以买到多少股
# 买入股票数是100的整数倍
Trade_count = round(
    (((float)(Trade_mount)) / ((float)(base_price))) / (GRID_COUNT * 100), 0) * 100
trade_count = (int)(Trade_count)

if (os.path.exists("op_plan.xlsx")):
    os.remove("op_plan.xlsx")

# 根据当日的收盘价，计算下一个交易日的涨／跌停价
#close = round((float)(Close), 2)
#close_up = close + round(close*10/100, 2)
#close_down = close - round(close*10/100, 2)
#print("close up %f, close down %f" % (close_up, close_down))

# 建立excel表格
wb = Workbook()
ws = wb.active
ws.title = "OP_PLAN"
ws.sheet_properties.tabColor = "ff0033"


def generate_base_grid():
    for i in range(1, GRID_COUNT + 1):
        if i <= UP_BASE_PRICE_GRID_COUNT:
            if i == 1:
                ws.cell(row=i, column=1).value = ((float)(price)) * \
                    ((100 + (int)(up_range)) / 100)
                ws.cell(row=i, column=2).value = ws.cell(
                    row=i, column=1).value * ((100 + (int)(up_range)) / 100)
            elif i == 2:
                ws.cell(row=i, column=1).value = ((float)(price))
                ws.cell(row=i, column=2).value = ws.cell(
                    row=i, column=1).value * ((100 + (int)(up_range)) / 100)
            else:
                ref_price = ws.cell(row=i - 1, column=1).value
                ws.cell(row=i, column=1).value = (
                    (float)(ref_price)) * ((100 - (int)(down_range)) / 100)
                ws.cell(row=i, column=2).value = (float)(ref_price)
            ws.cell(
                row=i,
                column=1).value = round(
                ws.cell(
                    row=i,
                    column=1).value,
                2)
            ws.cell(
                row=i,
                column=2).value = round(
                ws.cell(
                    row=i,
                    column=2).value,
                2)

            # 对低价股调整有意义，根据实际情况看是否需要调整
            ws.cell(row=i, column=3).value = trade_count - \
                round((UP_BASE_PRICE_GRID_COUNT - i) / 2, 0) * 100
            ws.cell(row=i, column=4).value = ws.cell(
                row=i, column=3).value * ws.cell(row=i, column=1).value
        else:
            ref_price = ws.cell(row=i - 1, column=1).value
            ws.cell(row=i, column=1).value = ((float)(ref_price)) * ((1000 - (int)
                                                                      (10 * down_range + (15 * (i - UP_BASE_PRICE_GRID_COUNT)))) / 1000)
            ws.cell(row=i, column=2).value = ws.cell(row=i, column=1).value * \
                ((10000 + (int)(up_range * 100 + (i - UP_BASE_PRICE_GRID_COUNT) * 40)) / 10000)
            ws.cell(
                row=i,
                column=1).value = round(
                ws.cell(
                    row=i,
                    column=1).value,
                2)
            ws.cell(
                row=i,
                column=2).value = round(
                ws.cell(
                    row=i,
                    column=2).value,
                2)

            # 对低价股调整有意义，根据实际情况看是否需要调整
            ws.cell(row=i, column=3).value = trade_count + \
                (i - UP_BASE_PRICE_GRID_COUNT) * 100
            ws.cell(row=i, column=4).value = ws.cell(
                row=i, column=3).value * ws.cell(row=i, column=1).value


def money_efficient():
    global Trade_mount
    Amount_money = 0
    Amount_stock = 0
    for i in range(1, GRID_COUNT + 1):
        Amount_money = Amount_money + ws.cell(row=i, column=4).value
        Amount_stock = Amount_stock + ws.cell(row=i, column=3).value

    Trade_mount = (float)(Trade_mount)
    if (Amount_money > Trade_mount):
        diff_money = Amount_money - Trade_mount
        for i in range(1, GRID_COUNT + 1):
            if (ws.cell(row=i, column=3).value == 0):
                continue
            tmp_money = ws.cell(row=i, column=1).value * 100

            ws.cell(
                row=i, column=3).value = ws.cell(
                row=i, column=3).value - 100
            ws.cell(row=i, column=4).value = ws.cell(
                row=i, column=3).value * ws.cell(row=i, column=1).value

            if (tmp_money < diff_money):
                diff_money = diff_money - tmp_money
            else:
                break

    if (Amount_money < Trade_mount):
        diff_money = Trade_mount - Amount_money
        for i in range(GRID_COUNT, 0, -1):
            tmp_money = ws.cell(row=i, column=1).value * 100

            if (tmp_money < diff_money):
                ws.cell(
                    row=i, column=3).value = ws.cell(
                    row=i, column=3).value + 100
                ws.cell(row=i, column=4).value = ws.cell(
                    row=i, column=3).value * ws.cell(row=i, column=1).value
                diff_money = diff_money - tmp_money
            else:
                break


def money_balance_among_grid():
    while True:
        money_high = 0
        money_low = 0
        high_index = 0
        low_index = 0
        for i in range(1, GRID_COUNT + 1):
            if (i == 1):
                money_high = ws.cell(row=i, column=4).value
                money_low = ws.cell(row=i, column=4).value
                high_index = i
                low_index = i
            else:
                if (money_high < ws.cell(row=i, column=4).value):
                    money_high = ws.cell(row=i, column=4).value
                    high_index = i
                if (money_low > ws.cell(row=i, column=4).value):
                    money_low = ws.cell(row=i, column=4).value
                    low_index = i

        if (money_high - money_low > ((float)(base_price) * 200)):
            ws.cell(
                row=low_index,
                column=3).value = ws.cell(
                row=low_index,
                column=3).value + 100
            ws.cell(row=low_index, column=4).value = ws.cell(
                row=low_index, column=3).value * ws.cell(row=low_index, column=1).value
            ws.cell(
                row=high_index,
                column=3).value = ws.cell(
                row=high_index,
                column=3).value - 100
            ws.cell(row=high_index, column=4).value = ws.cell(
                row=high_index, column=3).value * ws.cell(row=high_index, column=1).value
        else:
            break


def stock_op_plan():
    Amount_money = 0
    Amount_stock = 0
    for i in range(1, GRID_COUNT + 1):
        Amount_money = Amount_money + ws.cell(row=i, column=4).value
        Amount_stock = Amount_stock + ws.cell(row=i, column=3).value

    ws.cell(row=GRID_COUNT + 1, column=3).value = Amount_stock
    ws.cell(row=GRID_COUNT + 1, column=4).value = Amount_money

    for i in range(GRID_COUNT + 1, 0, -1):
        if i != GRID_COUNT + 1:
            ws.cell(row=i + 1, column=1).value = ws.cell(row=i, column=1).value
            ws.cell(row=i + 1, column=2).value = ws.cell(row=i, column=2).value
        ws.cell(row=i + 1, column=3).value = ws.cell(row=i, column=3).value
        ws.cell(row=i + 1, column=4).value = ws.cell(row=i, column=4).value

    ws.cell(row=1, column=1).value = 'Buy'
    ws.cell(row=1, column=2).value = 'Sell'
    ws.cell(row=1, column=3).value = 'Count'
    ws.cell(row=1, column=4).value = 'Money'

    ws.cell(row=GRID_COUNT + 3, column=3).value = init_mount
    ws.cell(row=GRID_COUNT + 3, column=4).value = Trade_mount
    ws.cell(row=GRID_COUNT + 3, column=2).value = loan

    wb.save("op_plan.xlsx")
    data_xls = pd.read_excel("op_plan.xlsx")
    print(data_xls)


if __name__ == '__main__':
    generate_base_grid()
    money_balance_among_grid()
    for i in range(1, 10):
        money_efficient()
    money_balance_among_grid()
    for i in range(1, 10):
        money_efficient()
    money_balance_among_grid()
    # for i in range(1, 2):
    #	money_efficient()
    stock_op_plan()

    print("Let's have fun!")
